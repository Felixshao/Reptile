import openpyxl, os, requests
from config.ProjectPath import get_project_path

path = get_project_path()
filepath = os.path.join(path, 'config', '爬虫接口配置文件.xlsx')
sheetname = 'Sheet1'


class ReadInter():

    def readExcel_config(self, filepath=filepath, sheetname=sheetname):
        """
        读取数据存入list中
        :param filepath: 文件路径
        :param sheetname: 表名称
        :return:
        """
        table = openpyxl.load_workbook(filepath)
        sheet = table[sheetname]

        sheets = []
        # 获取最大行列数
        row = sheet.max_row
        col = sheet.max_column
        for i in range(3, row+1):
            value_dict = {}
            for j in range(col):
                value_dict[sheet[2][j].value] = str(sheet[i][j].value).replace('’', "'").replace('‘', "'")
            sheets.append(value_dict)
        return sheets


if __name__ == '__main__':
    data = ReadInter().readExcel_config(os.path.join(path, 'config', '爬虫接口配置文件.xlsx'), 'Sheet1')[0]
    params = eval(data['参数'])
    print(data['请求头'], type(data['请求头']))
    headers = eval(data['请求头'])
    print(params, type(params))
    for i in range(1, 3):
        params['page'] = i
    if data['请求方式'] == 'get':
        t = requests.get(data['接口链接'], params=params, headers=headers)
        print(t.text)
